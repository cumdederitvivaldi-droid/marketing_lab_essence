# -*- coding: utf-8 -*-
import sys, io, os, glob, csv
from datetime import datetime
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

client = bigquery.Client(project='covering-app-ccd23')

FOLDER = "C:/Users/hound/OneDrive/바탕 화면/bigquery-agent/datas/crm_발송파일_26.04.22/새 폴더/"

# ── 1. 기존 A/B 그룹 CSV 로드 ───────────────────────────────────────────────
def load_user_ids(pattern):
    ids = set()
    for f in sorted(glob.glob(FOLDER + pattern)):
        with open(f, encoding='utf-8-sig') as fp:
            for row in csv.DictReader(fp):
                try:
                    ids.add(int(row['user_id']))
                except:
                    pass
    return ids

print("CSV 파일 로드 중...")
a_ids = load_user_ids("crm_A*.csv")
b_ids = load_user_ids("crm_B*.csv")
print(f"  기존 A그룹: {len(a_ids):,}명")
print(f"  기존 B그룹: {len(b_ids):,}명")

# ── 2. 쿠폰 198/199로 결제 완료한 유저 ─────────────────────────────────────
print("\n쿠폰 결제 완료 유저 조회 중...")
q_paid = """
SELECT DISTINCT ov.user_id, uc.coupon_policy_id AS policy_id
FROM `covering-app-ccd23.secure_dataset.order_v2` ov
JOIN `covering-app-ccd23.secure_dataset.user_coupon` uc ON ov.user_coupon_id = uc.id
WHERE uc.coupon_policy_id IN (198, 199)
  AND ov.status = 'COMPLETED'
  AND ov.deleted_at IS NULL
  AND ov.user_id IS NOT NULL
"""
paid_198 = set()
paid_199 = set()
for r in client.query(q_paid).result():
    if r.policy_id == 198:
        paid_198.add(r.user_id)
    else:
        paid_199.add(r.user_id)
print(f"  쿠폰 198 사용 완료: {len(paid_198):,}명")
print(f"  쿠폰 199 사용 완료: {len(paid_199):,}명")

# ── 3. 마지막 신청 이벤트 조회 (전체 유저 기준) ─────────────────────────────
print("\n마지막 신청 이벤트 조회 중...")
q_last = """
WITH order_product_type AS (
  SELECT ol.order_id,
    COUNTIF(p.product_code IN ('COVERING_BAG','LARGE_COVERING_BAG')) > 0 AS has_bag,
    COUNTIF(p.product_code LIKE 'PICKUP%') > 0 AS has_pickup
  FROM `covering-app-ccd23.secure_dataset.order_line` ol
  JOIN `covering-app-ccd23.secure_dataset.product` p ON ol.product_id = p.id
  WHERE ol.deleted_at IS NULL
  GROUP BY ol.order_id
),
last_order AS (
  SELECT ov.user_id,
    ov.id AS order_id,
    ov.status,
    ov.created_at,
    ROW_NUMBER() OVER (PARTITION BY ov.user_id ORDER BY ov.created_at DESC) AS rn
  FROM `covering-app-ccd23.secure_dataset.order_v2` ov
  WHERE ov.deleted_at IS NULL AND ov.user_id IS NOT NULL
)
SELECT lo.user_id,
  CASE
    WHEN opt.has_bag AND NOT opt.has_pickup THEN '봉투배송'
    WHEN opt.has_bag AND opt.has_pickup     THEN '수거+봉투'
    WHEN NOT opt.has_bag AND opt.has_pickup THEN '수거'
    ELSE '기타'
  END AS event_type,
  lo.status AS order_status,
  DATETIME(lo.created_at, 'Asia/Seoul') AS event_at
FROM last_order lo
LEFT JOIN order_product_type opt ON lo.order_id = opt.order_id
WHERE lo.rn = 1
"""
last_event = {}  # user_id -> (event_type, order_status, event_at)
for r in client.query(q_last).result():
    label = r.event_type + '(' + (r.order_status or '') + ')'
    last_event[r.user_id] = (label, str(r.event_at)[:19] if r.event_at else '')
print(f"  이벤트 조회 완료: {len(last_event):,}명")

# ── 4. 신규 A/B 그룹 조회 ─────────────────────────────────────────────────
print("\n신규 A/B 그룹 조회 중...")
q_new = """
WITH
-- 쿠폰 보유자 (active, not used)
coupon_holders AS (
  SELECT uc.user_id, uc.coupon_policy_id,
    uc.id AS user_coupon_id, uc.expire_date, uc.created_date AS issued_at
  FROM `covering-app-ccd23.secure_dataset.user_coupon` uc
  WHERE uc.coupon_policy_id IN (196, 197)
    AND uc.disabled_date IS NULL
    AND uc.deleted_date IS NULL
    AND (uc.expire_date IS NULL OR uc.expire_date > CURRENT_TIMESTAMP())
),
-- 서비스 지역 내 유저 (활성 user_address → active service_region)
in_service_area AS (
  SELECT DISTINCT ua.user_id
  FROM `covering-app-ccd23.secure_dataset.user_address` ua
  JOIN `covering-app-ccd23.secure_dataset.address` a ON ua.address_id = a.id
  JOIN `covering-app-ccd23.secure_dataset.service_region` sr ON a.h_code = sr.h_code
  WHERE ua.deleted_date IS NULL AND ua.active = true
    AND sr.active_flag = true AND sr.deleted_date IS NULL
),
-- 마케팅 동의 유저
marketing_agree AS (
  SELECT DISTINCT user_id
  FROM `covering-app-ccd23.secure_dataset.device`
  WHERE is_marketing_agree = true
),
-- 쿠폰 사용 완료 유저 (이미 사용함)
already_used AS (
  SELECT DISTINCT ov.user_id, uc.coupon_policy_id
  FROM `covering-app-ccd23.secure_dataset.order_v2` ov
  JOIN `covering-app-ccd23.secure_dataset.user_coupon` uc ON ov.user_coupon_id = uc.id
  WHERE uc.coupon_policy_id IN (196, 197)
    AND ov.status = 'COMPLETED'
    AND ov.deleted_at IS NULL
),
-- 전체 취소 유저 (주문이 있는데 전부 취소)
all_canceled AS (
  SELECT user_id
  FROM `covering-app-ccd23.secure_dataset.order_v2`
  WHERE deleted_at IS NULL AND user_id IS NOT NULL
  GROUP BY user_id
  HAVING COUNTIF(status != 'CANCELED') = 0
),
-- 봉투 배송 완료 유저 (BAG_ONLY COMPLETED)
home_kit_delivered AS (
  SELECT DISTINCT ov.user_id
  FROM `covering-app-ccd23.secure_dataset.order_v2` ov
  JOIN (
    SELECT ol.order_id,
      COUNTIF(p.product_code IN ('COVERING_BAG','LARGE_COVERING_BAG')) AS bag_cnt,
      COUNTIF(p.product_code LIKE 'PICKUP%') AS pickup_cnt
    FROM `covering-app-ccd23.secure_dataset.order_line` ol
    JOIN `covering-app-ccd23.secure_dataset.product` p ON ol.product_id = p.id
    WHERE ol.deleted_at IS NULL
    GROUP BY ol.order_id
  ) opt ON ov.id = opt.order_id
  WHERE ov.status = 'COMPLETED'
    AND ov.deleted_at IS NULL
    AND opt.bag_cnt > 0 AND opt.pickup_cnt = 0
    AND ov.user_id IS NOT NULL
),
-- 수거 전체 실패 유저
all_pickup_failed AS (
  SELECT ov.user_id
  FROM `covering-app-ccd23.secure_dataset.order_v2` ov
  JOIN `covering-app-ccd23.secure_dataset.fulfillment` f ON ov.id = f.order_id
  WHERE ov.deleted_at IS NULL AND ov.user_id IS NOT NULL
  GROUP BY ov.user_id
  HAVING COUNTIF(f.status = 'COMPLETED') = 0
    AND COUNTIF(f.status = 'FAILED') > 0
)
SELECT ch.user_id, ch.coupon_policy_id,
  ch.issued_at, ch.expire_date
FROM coupon_holders ch
JOIN in_service_area isa ON ch.user_id = isa.user_id
JOIN marketing_agree ma ON ch.user_id = ma.user_id
JOIN `covering-app-ccd23.secure_dataset.user` u ON ch.user_id = u.id
LEFT JOIN already_used au ON ch.user_id = au.user_id AND ch.coupon_policy_id = au.coupon_policy_id
LEFT JOIN all_canceled ac ON ch.user_id = ac.user_id
LEFT JOIN home_kit_delivered hk ON ch.user_id = hk.user_id
LEFT JOIN all_pickup_failed apf ON ch.user_id = apf.user_id
WHERE au.user_id IS NULL          -- 아직 사용 안 함
  AND u.withdrawal_date IS NULL    -- 탈퇴 안 함
  AND ac.user_id IS NULL           -- 전체 취소 아님
  AND hk.user_id IS NULL           -- 봉투 배송 완료 아님
  AND apf.user_id IS NULL          -- 수거 전체 실패 아님
ORDER BY ch.coupon_policy_id, ch.user_id
"""
new_a = []  # (user_id,)
new_b = []
for r in client.query(q_new).result():
    if r.coupon_policy_id == 196:
        new_a.append(r.user_id)
    else:
        new_b.append(r.user_id)
print(f"  신규 A그룹 (쿠폰 196): {len(new_a):,}명")
print(f"  신규 B그룹 (쿠폰 197): {len(new_b):,}명")

# ── 5. 기존 A/B 최종 대상 ─────────────────────────────────────────────────
final_a_existing = a_ids - paid_198
final_b_existing = b_ids - paid_199
print(f"\n기존 A그룹 최종: {len(final_a_existing):,}명 (제외: {len(paid_198 & a_ids):,}명)")
print(f"기존 B그룹 최종: {len(final_b_existing):,}명 (제외: {len(paid_199 & b_ids):,}명)")

# ── 6. Excel 생성 ─────────────────────────────────────────────────────────
print("\nExcel 생성 중...")

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL  = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
TOTAL_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="맑은 고딕", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")
thin = Side(style="thin", color="B8CCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ["user_id", "타겟그룹", "마지막신청이벤트종류", "이벤트시각"]
COL_W = [15, 25, 25, 22]

def build_sheet(ws, title_txt, rows_data):
    """rows_data: list of (user_id, group_label, event_type, event_at)"""
    ws.sheet_view.showGridLines = False

    # 제목
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = title_txt
    t.font = Font(name="맑은 고딕", bold=True, size=12, color="1F4E79")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # 헤더
    for ci, (col, w) in enumerate(zip(COLS, COL_W), 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # 데이터
    for ri, (uid, grp, etype, eat) in enumerate(rows_data, 3):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        vals = [uid, grp, etype, eat]
        aligns = [RIGHT, CENTER, CENTER, CENTER]
        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT; c.fill = fill; c.alignment = aln; c.border = BORDER
        ws.row_dimensions[ri].height = 18

    # 합계 행
    tr = len(rows_data) + 3
    for ci, (val, aln) in enumerate(zip([f"합계: {len(rows_data):,}명", "", "", ""], [LEFT,CENTER,CENTER,CENTER]), 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL; c.alignment = aln; c.border = BORDER
    ws.row_dimensions[tr].height = 22
    ws.freeze_panes = "A3"

def get_last(uid):
    if uid in last_event:
        return last_event[uid]
    return ("주문없음", "")

# ── 기존 A그룹
ws1 = wb.active
ws1.title = "기존A(지구의날)"
rows1 = sorted([(uid, "기존A(지구의날)", *get_last(uid)) for uid in final_a_existing], key=lambda x: x[0])
build_sheet(ws1, f"【기존 A그룹 — 지구의날 (쿠폰 EARTHDAY 미사용자)】  총 {len(rows1):,}명", rows1)

# ── 기존 B그룹
ws2 = wb.create_sheet("기존B(봄맞이)")
rows2 = sorted([(uid, "기존B(봄맞이)", *get_last(uid)) for uid in final_b_existing], key=lambda x: x[0])
build_sheet(ws2, f"【기존 B그룹 — 봄맞이 (쿠폰 SPRING 미사용자)】  총 {len(rows2):,}명", rows2)

# ── 신규 A그룹
ws3 = wb.create_sheet("신규A(지구의날)")
rows3 = [(uid, "신규A(지구의날)", *get_last(uid)) for uid in sorted(new_a)]
build_sheet(ws3, f"【신규 A그룹 — 지구의날 (쿠폰EARTH20 보유 활성 유저)】  총 {len(rows3):,}명", rows3)

# ── 신규 B그룹
ws4 = wb.create_sheet("신규B(봄맞이)")
rows4 = [(uid, "신규B(봄맞이)", *get_last(uid)) for uid in sorted(new_b)]
build_sheet(ws4, f"【신규 B그룹 — 봄맞이 (쿠폰SPRING20 보유 활성 유저)】  총 {len(rows4):,}명", rows4)

# ── 전체 합계 시트
ws5 = wb.create_sheet("전체요약")
ws5.sheet_view.showGridLines = False
ws5.merge_cells("A1:D1")
t5 = ws5["A1"]
t5.value = f"【CRM 2차 발송 타겟 요약 — 추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}】"
t5.font = Font(name="맑은 고딕", bold=True, size=12, color="1F4E79")
t5.alignment = CENTER
ws5.row_dimensions[1].height = 24

summary_headers = ["타겟그룹", "대상 유저 수", "설명"]
for ci, h in enumerate(summary_headers, 1):
    c = ws5.cell(row=2, column=ci, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
    ws5.column_dimensions[get_column_letter(ci)].width = [25, 18, 55][ci-1]
ws5.row_dimensions[2].height = 22

summary_rows = [
    ("기존A(지구의날)", len(rows1), f"1차 CRM A그룹 {len(a_ids):,}명 중 쿠폰198(EARTHDAY) 미사용 {len(paid_198 & a_ids):,}명 제외"),
    ("기존B(봄맞이)",   len(rows2), f"1차 CRM B그룹 {len(b_ids):,}명 중 쿠폰199(SPRING) 미사용 {len(paid_199 & b_ids):,}명 제외"),
    ("신규A(지구의날)", len(rows3), "쿠폰196(EARTH20) 보유·활성·서비스지역·마케팅동의·미사용·미탈퇴·봉투배송완료X"),
    ("신규B(봄맞이)",   len(rows4), "쿠폰197(SPRING20) 보유·활성·서비스지역·마케팅동의·미사용·미탈퇴·봉투배송완료X"),
    ("합계",           len(rows1)+len(rows2)+len(rows3)+len(rows4), "전체 2차 CRM 대상"),
]
for ri, (grp, cnt, desc) in enumerate(summary_rows, 3):
    fill = TOTAL_FILL if ri == len(summary_rows) + 2 else (ALT_FILL if ri % 2 == 0 else WHITE_FILL)
    font = TOTAL_FONT if ri == len(summary_rows) + 2 else BODY_FONT
    for ci, (val, aln) in enumerate(zip([grp, cnt, desc], [CENTER, RIGHT, LEFT]), 1):
        c = ws5.cell(row=ri, column=ci, value=val)
        c.font = font; c.fill = fill; c.alignment = aln; c.border = BORDER
        if ci == 2:
            c.number_format = "#,##0"
    ws5.row_dimensions[ri].height = 18

# ── 저장
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT = f"C:/Users/hound/OneDrive/바탕 화면/bigquery-agent/datas/{ts}_CRM_2차발송_타겟.xlsx"
wb.save(OUTPUT)
print(f"\n✅ 저장 완료: {OUTPUT}")
print(f"\n요약:")
print(f"  기존 A그룹: {len(rows1):,}명")
print(f"  기존 B그룹: {len(rows2):,}명")
print(f"  신규 A그룹: {len(rows3):,}명")
print(f"  신규 B그룹: {len(rows4):,}명")
print(f"  전체: {len(rows1)+len(rows2)+len(rows3)+len(rows4):,}명")
