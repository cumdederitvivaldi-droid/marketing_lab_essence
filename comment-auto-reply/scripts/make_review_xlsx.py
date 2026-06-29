# -*- coding: utf-8 -*-
"""
out/drafts.json → out/검토_답글.xlsx (팀원이 엑셀에서 검토/수정/승인).
컬럼: 번호 | 플랫폼 | 작성자 | 댓글 | 답글초안(수정가능) | 승인(O=게시) | 확인필요 | comment_id
- 자동 초안(확인필요 아님)은 승인칸에 'O'를 미리 채움. 확인필요 건은 비워둠(사람이 검토 후 O 입력).
"""
import sys, os, json
_usersite = os.path.join(os.environ.get("APPDATA", ""), "Python", "Python312", "site-packages")
if os.path.isdir(_usersite) and _usersite not in sys.path:
    sys.path.insert(0, _usersite)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "out"
DRAFTS = OUT / "drafts.json"
XLSX = OUT / "검토_답글.xlsx"

HEADERS = ["번호", "플랫폼", "작성자", "댓글", "답글초안(수정가능)", "승인(O=게시)", "확인필요", "comment_id"]


def main():
    if not DRAFTS.exists():
        raise SystemExit("out/drafts.json 이 없습니다. 먼저 generate_drafts.py 를 실행하세요.")
    drafts = json.loads(DRAFTS.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "답글검토"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, d in enumerate(drafts, 1):
        approve = "O" if (d["draft_reply"] and not d["needs_human"]) else ""
        ws.append([i, d["platform"], d["author"], d["comment_text"],
                   d["draft_reply"], approve, "예" if d["needs_human"] else "",
                   d["comment_id"]])
        if d["needs_human"]:
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=i + 1, column=c).fill = PatternFill("solid", fgColor="FFF2CC")

    widths = [5, 8, 16, 50, 55, 12, 9, 22]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    wb.save(XLSX)
    auto = sum(1 for d in drafts if d["draft_reply"] and not d["needs_human"])
    print(f"검토 파일 생성: {XLSX}")
    print(f"  총 {len(drafts)}건 / 자동승인(O) {auto}건 / 확인필요(노랑) {len(drafts) - auto}건")
    print("  엑셀에서 초안을 수정하고, 게시할 행의 '승인' 칸에 O 를 두세요. 저장 후 2_게시.bat 실행.")


if __name__ == "__main__":
    main()
